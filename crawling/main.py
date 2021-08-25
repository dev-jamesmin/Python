# coding=utf-8
# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

import requests
from bs4 import BeautifulSoup
from openpyxl.worksheet.worksheet import Worksheet
from requests.sessions import Session
import requests
import openpyxl

login_html = ''
params = dict()
params['loginId'] = ''
params['loginPwd'] = ''
params['mode'] = 'login'

modelId = []
fileName = 'result.xlsx'

totalCount = 1
pageCount = 2
onePageGoodsCount = 2

wb = openpyxl.Workbook()
sheet = wb.active


def get_html(url):
    _html = ""
    resp = requests.get(url)
    if resp.status_code == 200:
        _html = resp.text
    return _html


def setSession():
    login = session_data.post(login_html, params)
    login.raise_for_status()
    #  세션 설정을 위하여 최초한번 로드???
    loadPage("9800800056")


def loadPage(keywordValue):
    # keywordValue = "5012035901738"
    crawl_url = "" + str(keywordValue) + ""
    login = session_data.get(crawl_url)
    soupData = BeautifulSoup(login.content, 'html.parser')  # type: BeautifulSoup
    return soupData


def setExcelColumn():
    sheet.cell(1, 1).value = "번호"
    sheet.cell(1, 2).value = "상품명"
    sheet.cell(1, 3).value = "모델명"
    sheet.cell(1, 4).value = "판매여부"


# remove dummy data type-2
def getTitle(originResult):
    arrMid = originResult.split(" title=")
    arrLast = arrMid[1].split(" width=")
    return arrLast[0]


def getData(originResult):
    arr = str(originResult).split('item-display type-gallery')
    # print(arr[1])
    arrMid = arr[1].split("<strong>")
    # print(arrMid)
    arrLast = arrMid[1].split("</strong>")
    # print(arrLast[0])
    return arrLast[0]


def ExcelRead():
    global modelId
    global totalCount
    workbook = openpyxl.load_workbook('list.xlsx')
    ws = workbook['name']

    for cell in ws['G']:  # A열의 모든 셀을 확인
        # print(str(cell.value).strip())
        typeCase = "0"
        devalue = str(cell.value).strip()
        if devalue == "None":
            # print("is None")
            typeCase = "1"
        elif devalue == "모델명":
            # print("is 모델명")
            typeCase = "2"
        else:
            # print(str(cell.value).strip())
            modelId.append(devalue)

    return modelId


def dataArray(soupData):
    galleryData = soupData.find('div', class_='item-display type-gallery')
    thumbnailData = galleryData.find_all('div', class_='thumbnail')
    return thumbnailData


def excelWrite(keywordValue, resultData):
    global totalCount
    checkWord = "soldout-img"

    print(totalCount)
    print(resultData)
    print(len(resultData))
    # testData.
    # href.find("a")["href"]
    # 파싱 방법 찾음.
    print(resultData.find("img")["title"])

    sheetNumBer = totalCount + 1
    sheet.cell(sheetNumBer, 1).value = str(totalCount)  # "번호"
    sheet.cell(sheetNumBer, 2).value = str(getTitle(str(resultData)))
    sheet.cell(sheetNumBer, 3).value = str(keywordValue)

    if checkWord in str(resultData):
        sheet.cell(sheetNumBer, 4).value = str("품절")
    else:
        sheet.cell(sheetNumBer, 4).value = str("판매중")
    wb.save(fileName)
    totalCount += 1


def checkData(keywordValue):
    global totalCount

    soupData = loadPage(keywordValue)
    resultArray = dataArray(soupData)  # 검색 결과 배열

    j = 0
    while j < len(resultArray):
        excelWrite(keywordValue, resultArray[j])
        j += 1


if __name__ == '__main__':
    session_data = requests.session()  # type: Session
    totalCount = 1
    print("[START]")
    setSession()
    modelId = ExcelRead()

    wb = openpyxl.Workbook()
    sheet = wb.active
    setExcelColumn()

    i = 0
    while i < len(modelId):
        checkData(str(modelId[i]))
        # wb.save(fileName)
        i += 1

    print('[END]')
